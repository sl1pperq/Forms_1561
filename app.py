import qrcode
from flask import *
import json
import random
from openpyxl import Workbook

app = Flask(__name__)
app.secret_key = 'school'

events = []
users = []

try:
    with open('json/events.json', 'r') as file:
        events = json.loads(file.read())
    with open('json/users.json', 'r') as file:
        users = json.loads(file.read())
except Exception as e:
    print(e)


def save_data():
    with open('json/events.json', 'w') as file:
        file.write(json.dumps(events, ensure_ascii=False))
    with open('json/users.json', 'w') as file:
        file.write(json.dumps(users, ensure_ascii=False))


@app.route('/')
def main_page():
    if session.get('auth', False) == False:
        return render_template('main.html', auth=0)
    else:
        return render_template('main.html', auth=1, events=events)


@app.route('/signup', methods=['POST'])
def signup_page():
    name = request.form.get("name")
    fami = request.form.get("famil")
    otch = request.form.get("otch")
    email = request.form.get("email")
    password = request.form.get("password")
    users.append({
        "fami": fami,
        "name": name,
        "otch": otch,
        "email": email,
        "password": password
    })
    save_data()
    session['auth'] = email
    return redirect('/')


@app.route('/login', methods=['POST'])
def login_page():
    email = request.form.get("email")
    password = request.form.get("password")
    for user in users:
        if user['email'] == email:
            if user['password'] == password:
                session['auth'] = email
                return redirect('/')
    return redirect('/')


@app.route('/create', methods=['GET'])
def get_create():
    return render_template("create.html")


@app.route('/create', methods=['POST'])
def post_create():
    name = request.form.get("name")
    description = request.form.get("description")

    studfio = request.form.get("studfio")
    parfio = request.form.get("parfio")

    studmail = request.form.get("studmail")
    parmail = request.form.get("parmail")

    studphone = request.form.get("studphone")
    parphone = request.form.get("parphone")

    clas = request.form.get("class")
    school = request.form.get("school")
    comment = request.form.get("comment")

    id = str(random.randint(100000, 999999))
    events.append({
        'id': id,
        'name': name,
        'description': description,
        'studfio': studfio,
        'parfio': parfio,
        'parmail': parmail,
        'studmail': studmail,
        'parphone': parphone,
        'studphone': parphone,
        'clas': clas,
        'school': school,
        'comment': comment,
        'answers': []
    })
    save_data()
    return redirect('/')


@app.route('/event/solve/<id>')
def event(id):
    for event in events:
        if event['id'] == id:
            return render_template("event.html", event=event)
    return redirect('/')


@app.route('/event/solve/<id>', methods=['POST'])
def solve_event(id):
    user_answers = {}

    try:
        description = request.form.get("description")
        user_answers['description'] = description
    except:
        pass

    try:
        studfio = request.form.get("studfio")
        user_answers['studfio'] = studfio
    except:
        pass

    try:
        parfio = request.form.get("parfio")
        user_answers['parfio'] = parfio
    except:
        pass

    try:
        studmail = request.form.get("studmail")
        user_answers['studmail'] = studmail
    except:
        pass

    try:
        parmail = request.form.get("parmail")
        user_answers['parmail'] = parmail
    except:
        pass

    try:
        studphone = request.form.get("studphone")
        user_answers['studphone'] = studphone
    except:
        pass

    try:
        parphone = request.form.get("parphone")
        user_answers['parphone'] = parphone
    except:
        pass

    try:
        clas = request.form.get("class")
        user_answers['clas'] = clas
    except:
        pass

    try:
        school = request.form.get("school")
        user_answers['school'] = school
    except:
        pass

    try:
        comment = request.form.get("comment")
        user_answers['comment'] = comment
    except:
        pass

    print(user_answers)
    for event in events:
        if event['id'] == id:
            event['answers'].append(user_answers)
            save_data()
            return render_template('success.html')
    return redirect('/')

@app.route('/event/result/<id>')
def result_event(id):
    for event in events:
        if event['id'] == id:
            return render_template("result.html", answers=event['answers'], id=id)
    return redirect('/')

@app.route('/event/delete/<id>')
def delete_event(id):
    for e in range(len(events)):
        if events[e]['id'] == id:
            del events[e]
            save_data()
            break
    return redirect('/')

@app.route('/event/export/xlsx/<id>')
def event_export(id):
    for event in events:
        if event['id'] == id:
            with open('file.xlsx', 'w') as file:
                file = Workbook()
                ws = file.active
                alphabet = ['A', 'B', 'C', 'D', 'E',' F', 'G', 'H', 'I']
                ws['A1'] = 'ФИО ученика'
                ws['B1'] = 'ФИО родителя'
                ws['C1'] = 'Почта ученика'
                ws['D1'] = 'Почта родителя'
                ws['E1'] = 'Телефон ученика'
                ws['F1'] = 'Телефон родителя'
                ws['G1'] = 'Класс'
                ws['H1'] = 'Школа'
                ws['I1'] = 'Комментарий'
                for a in alphabet:
                    for i in range(len(event['answers'])):
                        if a == 'A':
                            ws[f'{a}{i + 2}'] = event['answers'][i]['studfio']
                        elif a == 'B':
                            ws[f'{a}{i + 2}'] = event['answers'][i]['parfio']
                        elif a == 'C':
                            ws[f'{a}{i + 2}'] = event['answers'][i]['studmail']
                        elif a == 'D':
                            ws[f'{a}{i + 2}'] = event['answers'][i]['parmail']
                        elif a == 'E':
                            ws[f'{a}{i + 2}'] = event['answers'][i]['studphone']
                        elif a == 'F':
                            ws[f'{a}{i + 2}'] = event['answers'][i]['parphone']
                        elif a == 'G':
                            ws[f'{a}{i + 2}'] = event['answers'][i]['clas']
                        elif a == 'H':
                            ws[f'{a}{i + 2}'] = event['answers'][i]['school']
                        elif a == 'I':
                            ws[f'{a}{i + 2}'] = event['answers'][i]['comment']
                file.save('./file.xlsx')
                return send_file('file.xlsx')

@app.route('/event/create/qr/<id>')
def qr_event_create(id):
    with open('qr.png', 'w') as file:
        file = qrcode.make(f'http://127.0.0.1:5000/event/solve/{id}')
        file.save('./qr.png')
        return send_file('qr.png')

app.run()
